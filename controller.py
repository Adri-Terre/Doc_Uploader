import vista
import os
import shutil       

from tkinter.messagebox import *
from tkinter import filedialog
from tkinter import *
from tkcalendar import Calendar



def ruta():
    folder_selected = filedialog.askdirectory()
    return folder_selected

def file():
    import module_variable as mod_var
    file_selected = filedialog.askopenfilename()
    fichero = open(file_selected,'r')
    mod_var.file_origen = file_selected
    return file_selected

def upload():
    import psutil
    from module_variable import file_origen, ruta_destino,aeropuertos
    from module_variable import regex_numero,fecha_seteada
    import os, re, errno
    import module_variable as mod_var
    from tkinter import ttk 

    aep_seleccionado= vista.combo_aep.get()
    sistema_seleccionado = vista.combo_sistema.get()
    folio_seleccionado = vista.combo_folio.get()
    fir_seleccionada= vista.combo_anexo.get()
    re_match = False


    if file_origen != "": 
        a= file_origen.rindex("/")
        ruta_origen = file_origen[:(a+1)]
        b = file_origen.rindex(".")
        extension_archivo = file_origen[(b):] 

    else:
        showinfo(
            "Campo incompleto",
            "Asigne la ruta origen",)
        
    try:
        if (vista.w9.winfo_ismapped()) == True:
            vista.w9.destroy()
    except:
           print ("label exception")

    if re.match(regex_numero, vista.folio_input.get()):
        re_match = True
    

    if (fecha_seteada != "" and aep_seleccionado != "" and folio_seleccionado != "" and vista.folio_input.get() != "")  or (fecha_seteada != "" and mod_var.checkbox.get() == True):
        if (re_match  == True) or mod_var.checkbox.get() == True:         
            
            if (folio_seleccionado== "CHECK LIST I") or (folio_seleccionado== "CHECK LIST II"): 
                nombre_check = str(fecha_seteada + " " + aep_seleccionado + " " + sistema_seleccionado + " " + folio_seleccionado + " " + vista.folio_input.get() + extension_archivo)

                archivo_renombrado = os.path.join (ruta_origen,nombre_check)#ruta_origen + nombre_nuevo
                
            elif mod_var.checkbox.get() == True:
                a= fecha_seteada.rindex(".")
                fecha_mes = fecha_seteada[:(a)]
                nombre_anexo = str(fecha_mes + " " + "("+ aep_seleccionado + ")" +  " " + fir_seleccionada + "" + " anexo VIII" + "" + extension_archivo)
                archivo_renombrado = os.path.join (ruta_origen,nombre_anexo)
            else:    
                nombre_nuevo = str(fecha_seteada + " " + aep_seleccionado + " " + sistema_seleccionado + " " + folio_seleccionado + " " + vista.folio_input.get() + extension_archivo)
        

                archivo_renombrado = os.path.join (ruta_origen,nombre_nuevo)#ruta_origen + nombre_nuevo

            c= fecha_seteada.index(".")
            año_seteado = fecha_seteada[:(c)]

           
            if (folio_seleccionado== "CHECK LIST I") or (folio_seleccionado== "CHECK LIST II"):    
                    try:
                        nueva_carpeta = ruta_destino + "/" + aep_seleccionado + "/" + año_seteado + "/" + sistema_seleccionado + "/" + "Check list"
                        os.mkdir(nueva_carpeta)
                        
                    except OSError as e:
                        if e.errno == errno.EEXIST:
                            print ("La carpeta ya existe")
                    
                    ruta_destino2 = ruta_destino + "/" + aep_seleccionado + "/" + año_seteado + "/" + sistema_seleccionado + "/" + "Check list" + "/" + nombre_check
            
            elif mod_var.checkbox.get() == True:
                    try:
                            nueva_carpeta = ruta_destino + "/" + "z.Anexos VIII" + "/" + año_seteado + "/" + fecha_mes
                            os.mkdir(nueva_carpeta)
                    except OSError as e:
                        if e.errno == errno.EEXIST:
                            print ("La carpeta ya existe")
                    try:
                            nueva_carpeta = ruta_destino + "/" + "z.Anexos VIII" + "/" + año_seteado + "/" + fecha_mes + "/" + fir_seleccionada
                            os.mkdir(nueva_carpeta)
                    except OSError as e:
                        if e.errno == errno.EEXIST:
                            print ("La carpeta ya existe")
                    
                    ruta_destino2 = ruta_destino + "/" + "z.Anexos VIII" + "/" + año_seteado + "/" + fecha_mes + "/" + fir_seleccionada + "/" + nombre_anexo
                    folio_seleccionado = "Anexo VIII " + fir_seleccionada
            else:
                    ruta_destino2 = ruta_destino + "/" + aep_seleccionado + "/" + año_seteado + "/" + sistema_seleccionado + "/" + nombre_nuevo
                    
            
            try:
                
                match extension_archivo:

                    case ".pdf":
                        proc= "Acrobat.exe" in (i.name() for i in psutil.process_iter()) 
                        if proc == True:
                            os.system(f"taskkill /im Acrobat.exe /f")
                    case ".jpg":
                        proc= "PhotosApp.exe" in (i.name() for i in psutil.process_iter()) 
                        if proc == True:
                            os.system(f"taskkill /im PhotosApp.exe /f")
                    case ".png":
                        proc= "PhotosApp.exe" in (i.name() for i in psutil.process_iter()) 
                        if proc == True:
                            os.system(f"taskkill /im PhotosApp.exe /f")
                    case ".xslx":  
                        proc= "EXCEL.EXE" in (i.name() for i in psutil.process_iter()) 
                        if proc == True:
                            os.system(f"taskkill /im EXCEL.EXE /f")
                    case ".doc":            
                        proc= "WINWORD.EXE" in (i.name() for i in psutil.process_iter()) 
                        if proc == True:
                            os.system(f"taskkill /im WINWORD.EXE /f")        
                """
                if (extension_archivo == ".pdf"):
                    proc= "Acrobat.exe" in (i.name() for i in psutil.process_iter()) 
                    if proc == True:
                        os.system(f"taskkill /im Acrobat.exe /f")
                
                if (extension_archivo == ".jpg")or (extension_archivo == ".png")or(extension_archivo == ".tiff"):
                    proc= "PhotosApp.exe" in (i.name() for i in psutil.process_iter()) 
                    if proc == True:
                        os.system(f"taskkill /im PhotosApp.exe /f")
                
                if (extension_archivo == ".xslx"):
                    proc= "EXCEL.EXE" in (i.name() for i in psutil.process_iter()) 
                    if proc == True:
                        os.system(f"taskkill /im EXCEL.EXE /f")
                
                if (extension_archivo == ".doc"):
                    proc= "WINWORD.EXE" in (i.name() for i in psutil.process_iter()) 
                    if proc == True:
                        os.system(f"taskkill /im WINWORD.EXE /f")
                """
                
                os.rename(file_origen, archivo_renombrado)
                
                print ("Se reemplaza el archivo: " + str(file_origen) + "por: "+str(archivo_renombrado))

                
                shutil.move(archivo_renombrado,ruta_destino2)           
                    
                vista.w9 = vista.Label(vista.master, text= ruta_destino2, foreground="green")
                vista.w9.place(x=10, y=330)
                vista.contador_archivos_cargados = vista.contador_archivos_cargados + 1
                text = str (vista.contador_archivos_cargados)
                vista.w2.destroy()
                w2 = vista.Label(vista.master, text= text, foreground="black")
                w2.place(x=820, y=310)

                vista.w13.destroy()
                print(mod_var.checkbox_falla)
                
                if (mod_var.checkbox_falla.get() == True):
                    falla = "Corresponde"
                else:
                    falla = ""
                """
                excel(aep_seleccionado,sistema_seleccionado,folio_seleccionado,vista.folio_input.get(),fecha_seteada,año_seteado,archivo_renombrado,falla)

                vista.folio_input.delete(0,END)
                vista.w8.destroy()   
                vista.w7.destroy()                   
                vista.combo_aep.set("")
                vista.combo_sistema.set("")
                vista.combo_folio.set("")
                vista.combo_anexo.set("")
                vista.mod_var.checkbox.set(0)
                vista.mod_var.checkbox_falla.set(0)

                """    
                



            except:
                showinfo("Advertencia", "Cierre el archivo seleccionado")
                log_rechazado(archivo_renombrado, "Al intentar subir el archivo al drive")
        else:
            showinfo("Mensaje Folio", "El período ingresado deben ser números")
    else:
        showinfo(
            "Campos incompletos",
            "Complete todos los campos",)
    
    
def excel(airport,aid,attachment,number,date,año_seteado,archivo_para_cargar,falla):
    from datetime import datetime
    import xlsxwriter
    from openpyxl import load_workbook
    import openpyxl
    import module_variable as mod_var
    import os, errno

    try:

        carpeta_excel = mod_var.ruta_destino_excel + "/" + año_seteado
        os.mkdir(carpeta_excel)

    except OSError as e:
        if e.errno == errno.EEXIST:
            print ("La carpeta ya existe")
    
    ruta_excel_online = carpeta_excel + "/" + "DNAV - Checklist 1 - Anexos.xlsx"
    try:    
        
        
        if not os.path.exists(ruta_excel_online):
            workbook = xlsxwriter.Workbook(ruta_excel_online)
            worksheet = workbook.add_worksheet()
            column = 0
            row = 2
            content = ("AEROPUERTO","RADIOAYUDA","DOCUMENTO","FOLIO","FECHA","FALLA")
            
            for item in content :

                # write operation perform
                worksheet.write(row, column, item)

                # incrementing the value of row by one
                # with each iterations.
                column += 1
            
            img = openpyxl.drawing.image.Image('img_miniatura.png')

            workbook.close()

            wb = load_workbook(ruta_excel_online)
            ws = wb.active
            ws.add_image(img, 'A1')

            wb.save(ruta_excel_online)

            #workbook.close()
          
        wb_row = load_workbook(ruta_excel_online)
        sheet = wb_row["Sheet1"]
        wb_add = load_workbook(ruta_excel_online)

        sheet = wb_add["Sheet1"]
        
        #convert returned last row to string
        print("Last Row with empty rows by using len: ", len(sheet['B']))
        ins_row = str(len(sheet['B']) + 1 )
        
        date = date.replace(".","/")
        date = datetime.strptime(date, '%Y/%m/%d')
        sheet["A"+ins_row] = airport
        sheet["B"+ins_row] = aid
        sheet["C"+ins_row] = attachment
        sheet["D"+ins_row] = number
        sheet["E"+ins_row] = date
        sheet["F"+ins_row] = falla
        
        #Save data in the Workbook
        
        wb_add.save(ruta_excel_online)
            # iterating through content list
        
        vista.w13 = vista.Label(vista.master, text= "OK", foreground="green")
        vista.w13.place(x=180, y=360)

        vista.contador_registros_cargados = vista.contador_registros_cargados + 1
        text = str (vista.contador_registros_cargados)
        vista.w15.destroy()
        w15 = vista.Label(vista.master, text= text, foreground="black")
        w15.place(x=820, y=360)
    
    except:
        showinfo("Advertencia", "File Excel Error")
        log_rechazado(archivo_para_cargar, "Al cargar el registro en el DNAV Checklist I")
        vista.w13 = vista.Label(vista.master, text= "NO", foreground="red")
        vista.w13.place(x=180, y=360)

def not_in_use(filename):
        try:
            os.rename(filename,filename)
            return True
        except:    
            return False
#if not_in_use(excelFilePath):
#wb.save(excelFilePath)
     

def log_rechazado(archivo_rechazado,tipo):
    import logging
    """esta función genera un log con los errores y archivos rechazados"""

    
    try:

        if len(archivo_rechazado) != 0:

            logging.basicConfig(
                filename="Archivos_rechazados_" + ".log",
                filemode="a",
                format="%(asctime)s : %(levelname)s : %(message)s",
                datefmt="%d/%m/%y %H:%M:%S",
                level=logging.INFO,
            )
            a = 0

            #for x in archivo_rechazado:

            logging.info("Se ha Rechazado el archivo: " + str(archivo_rechazado) + " -> " + str(tipo) + "\n"
            )
            a += 1

            logging.shutdown()

        showinfo("Advertencia", "Se ha generado un log")

    except:
        print("Error inesperado")
        logging.fatal("Error inesperado")
